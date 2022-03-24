import pandas as pd
import os,sys
from typing import Tuple, List, Any, AnyStr

def add_piece_of_decoy_name_as_subset( df: pd.DataFrame, position: int, new_column_name: str)-> pd.DataFrame:
    """
    This takes a dataframe that has a 'decoy' column, split by "_". Takes the position as a new subset.
    Useful where experiments are contained in decoy names. Position can be negative.
    """
    def get_new_column(s):
        x = s.split('_')
        return x[position]

    if not 'decoy' in df.columns:
        sys.exit("No decoy field in column!  Cannot add subset based on decoy name!")

    df[new_column_name] = df['decoy'].apply( get_new_column )
    return df

def calc_means(d, column="rmsd") -> Tuple[float, float]:
    """
    Return a tuple of mean/std
    """
    avg_rmsd = d[column].mean()
    sd = d[column].std()
    return tuple(avg_rmsd, sd)

def rank_data(d, rank_on='total_score', new_column='score_rank') -> pd.DataFrame:
    """
    Rank data and insert into DF as a new column.
    :param d:
    :param rank_on:
    :param new_column:
    :return:
    """
    d2 = d.set_index(['total_score']).sort_index().reset_index()
    d2[new_column] = d2.index + 1
    return d2

def multi_tab_excel(df_list, sheet_list, file_name):
    """
    Writes multiple dataframes as separate sheets in an output excel file.

    If directory of output does not exist, it will create it.

    Author: Tom Dobbs
    http://stackoverflow.com/questions/32957441/putting-many-python-pandas-dataframes-to-one-excel-worksheet


    :param df_list: [pd.Dataframe]
    :param sheet_list: [str]
    :param file_name: str

    """
    if not os.path.exists(os.path.dirname(file_name)):
        os.mkdir(os.path.dirname(file_name))

    writer = pd.ExcelWriter(file_name,engine='xlsxwriter')
    for dataframe, sheet in zip(df_list, sheet_list):
        dataframe.to_excel(writer, sheet_name=sheet, startrow=0 , startcol=0)
    writer.save()


def add_sheet_to_workbook(local_wb, local_df, name):
    """
    Adds a Dataframe as a sheet to an openpyxl workbook using dataframe_to_rows
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    ws = local_wb.create_sheet(name)
    ws.title = name
    for r in dataframe_to_rows(local_df, index=False, header=True):
        ws.append(r)

    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'

def drop_matching(local_df: pd.DataFrame, column: str, datapoint: str) -> pd.DataFrame:
    """
    Drop a column, return a new df.
    :param df:
    :param column:
    :return:
    """
    local_df = local_df[local_df[column] != datapoint]
    return local_df

def drop_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Drop Duplicate columns from the DataFrame.
    Return DF

    :param df: dataframe.DataFrame
    :rtype: pandas.DataFrame
    """
    return df.T.groupby(level=0).first().T

def detect_numeric(df: pd.DataFrame) -> pd.DataFrame:
    """
    Detect numeric components

    :param df: pd.DataFrame
    :rtype: pd.DataFrame

    """
    #return df.convert_objects(convert_numeric=True)
    return df.infer_objects()

def get_columns(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    """
    Get a new dataframe of only the columns

    :param df: dataframe.DataFrame
    :param columns: list
    :rtype: pd.DataFrame
    """
    return df[columns]

def get_matches(df: pd.DataFrame, column: str, to_match: str) -> pd.DataFrame:
    """
    Get all the rows that match a paricular element of a column.

    :param df: dataframe.DataFrame
    :param column: str
    :param to_match: str
    :rtype: pd.DataFrame
    """

    return df[df[column] == to_match]

def get_multiple_matches(df: pd.DataFrame, column: str, to_match_array: List[str]) -> pd.DataFrame:
    """
    Get all the rows that match any of the values in to_match_array.

    :param df: dataframe.DataFrame
    :param column: str
    :param to_match_array: list
    :rtype: pd.DataFrame
    """
    return df[df[column].isin(to_match_array)]

def get_match_by_array(df: pd.DataFrame, column: str, match_array: pd.Series) -> pd.DataFrame:
    """
    Get a new dataframe of all dataframes of the subset series, match_array

    Note: This will result in a dataframe, but there may be strange issues when you go to plot the data in seaborn
            No idea why.

    :param df: pd.DataFrame
    :param column: str
    :param match_array: pd.Series
    :rtype: pd.DataFrame
    """

    new_df = df[df[column].isin(match_array)]
    return new_df

def get_row_matches(df: pd.DataFrame, column1: str, to_match:str, column2) -> pd.Series:
    """
    Get the elements of the rows that match a particular column.  If one element, this can be converted easily enough
    :param df: pd.DataFrame
    :param column1: str
    :param to_match: str
    :param column2: str
    :rtype: pd.Series
    """

    return df[df[column1] == to_match][column2]

def get_value(df: pd.DataFrame, column: str) -> Any:
    """
    Get a single value from a one-row df.  THis is to help for implicit docs, since the syntax to Iloc is so fucking strange.

    :param df: pd.DataFrame
    :param column: str
    :return: value
    """
    return df.iloc[0][column]

def get_n_matches(df: pd.DataFrame, column: str, to_match: str) -> int:
    """
    Get the number of matches
    :param df: pd.DataFrame
    :param column: str
    :param to_match:
    :rtype: int 
    """
    return len(get_matches(df, column, to_match))

def sort_on_list(df: pd.DataFrame, column: AnyStr, sort_order: List[str]) -> pd.DataFrame:
    """
    Given a list of values, and a column, create a new dataframe that is sorted like so. 
    No idea why this is so difficult.
    :param df: 
    :param list_to_sort: 
    :rtype: pd.DataFrame 
    """
    # Sort:
    sep = []
    for o in sort_order:
        sep.append(df[df[column].isin([o])])
    return pd.concat(sep).reset_index(drop=True)



class DataFrame2(pd.DataFrame):
    """
    An attempt at sublcassing a dataframe dataframe
    """
    def __init__(self, data=None, index=None, columns=None, dtype=None,copy=False):
        pd.DataFrame.__init__(data=data, index=index, columns=columns, dtype=dtype,copy=copy)

    def drop_duplicate_columns(self):
        """
        Drop Duplicate columns from the DataFrame in place
        :return:
        """

        #I'm not sure how to do this inplace, without reassigning self.  If you know, please edit this.


        self = self.T.groupby(level=0).first().T

    def detect_numeric(self):
        self = self.infer_objects()

    def get_columns(self, columns):
        return self[columns]

    def get_matches(self, column, to_match):
        """
        Get all the rows that match a paricular element of a column.
        :param column: str
        :param to_match: str
        :rtype: pandas.DataFrame
        """

        return self[self[column] == to_match]

    def get_row_matches(self, column1, to_match, column2):
        """
        Get the elements of the rows that match a particular column.  If one element, this can be converted easily enough
        :param column1: str
        :param to_match: str
        :param column2: str
        :rtype: pandas.Series
        """

        return self[self[column1] == to_match][column2]

    def n_matches(self, column, to_match):
        """
        Return the number of matches.
        :param column: str
        :param to_match: str
        :rtype: int
        """
        return len(get_matches(column, to_match))

    def to_tsv(self, path_or_buf=None, na_rep='', float_format=None,
               columns=None, header=True, index=True, index_label=None,
               mode='w', encoding=None, compression=None, quoting=None,
               quotechar='"', line_terminator='\n', chunksize=None,
               tupleize_cols=False, date_format=None, doublequote=True,
               escapechar=None, decimal='.'):
        self.to_csv(sep = "\t", path_or_buf=path_or_buf, na_rep=na_rep, float_format=float_format,
               columns=columns, header=header, index=index, index_label=index_label,
               mode=mode, encoding=encoding, compression=compression, quoting=quoting,
               quotechar=quotechar, line_terminator=line_terminator, chunksize=chunksize,
               tupleize_cols=tupleize_cols, date_format=date_format, doublequote=doublequote,
               escapechar=escapechar, decimal=decimal)

#Notes
    #To return a DF instead of a damn group by object:
        #for name, dd in residues.groupby('PDB ID', as_index = False):

